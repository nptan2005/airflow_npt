import os
import re
import inspect
import pandas as pd
from openpyxl import load_workbook

from . import DataConstant
from core_config import import_configuration


class FileImport:
    def __init__(self,fileName:str, templateName:str):

        self._template = import_configuration.import_template[templateName]
        self._fileName = fileName
        self._sep = self._template.separate
        
        
        # ########
        self._note = ''


    def __enter__(self):
        return self

    def __del__(self):
        self._clear_attributes()	
        

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._clear_attributes()

    def _clear_attributes(self):
        """Clear object attributes during exit."""
        attributes = [
            '_fileName', '_note', '_path_key', 
            '_template'
            ]
        for attr in attributes:
            setattr(self, attr, None)

    @property
    def note(self) -> str:
        return self._note
    
    @property
    def template(self):
        return self._template
      

    @property
    def fileName(self):
        return self._fileName
    


    def configEndRow(self, maxRow: int) -> int:
        end_row = self.template.end_row
        
        if end_row <= 0:
            end_row = maxRow

        return end_row


    def configEndCol(self, maxCol: int) -> int:
        end_col = self.template.end_col
        
        if self.template.end_col <= 0:
            end_col = maxCol

        self._endColumnNumer = end_col
        return end_col
    
    @property
    def endColumnNumer(self):
        return self._endColumnNumer



    
    
    def mapDataTypeFromPandasToOraclDType(self, df: pd.DataFrame) -> list:
        params = []
        for col in df.columns:
            params.append(self.mapDataType(df[col].dtype))

        return params
    # end mapDataTypeFromPandasToOraclDType
    def __check_max_value(self, value, max_value, min_value=None):
        if value > max_value or (min_value is not None and value < min_value):
            raise ValueError(f"Value {value} out of bounds for column: [{min_value}, {max_value}]")
        return value
    # end __check_max_value

    def prepare_oracle_df(self, df_schema: pd.DataFrame, dfImp: pd.DataFrame) -> pd.DataFrame:

        dtype_conversion_map = {
            ('VARCHAR2', 'datetime64'): lambda df_col,precision, scale,max_length: df_col.dt.strftime('%Y%m%d%H%M%S').str.slice(0, max_length),
            ('DATE', 'numeric'): lambda df_col,precision, scale,max_length: self.__convert_to_datetime(df_col),
            ('DATE', 'str'): lambda df_col,precision, scale,max_length: self.__convert_to_datetime(df_col),
            ('VARCHAR2', 'str'): lambda df_col,precision, scale,max_length: df_col.astype(str).fillna('').str.slice(0, max_length),
            ('VARCHAR2', 'bool'): lambda df_col,precision, scale,max_length: df_col.astype('str', errors='ignore'),
            ('NUMBER', 'bool'): lambda df_col,precision, scale,max_length: self._convert_bool_to_value(df_col.fillna('0'), 'NUMBER'),
            ('VARCHAR2', 'numeric'): lambda df_col,precision, scale,max_length: self.__convert_number_to_str(df_col,max_length),
            ('NUMBER', 'numeric'): lambda df_col,precision, scale,max_length: self.__convert_number_dtype(df_col,precision,scale),
            ('NUMBER', 'str'): lambda df_col,precision, scale,max_length: self.convert_str_to_number(df_col,precision,scale),
            
        }
        
        for index, row in df_schema.iterrows():
            column = row['COLUMN_NAME']
            if column not in dfImp.columns:
                continue

            dbDataType = row['DATA_TYPE']
            max_length = int(row['MAX_LENGTH']) if not pd.isna(row['MAX_LENGTH']) else None
            precision = int(row['PRECISION']) if not pd.isna(row['PRECISION']) else None
            scale = int(row['SCALE']) if not pd.isna(row['SCALE']) else None
            dfColumnDtype = dfImp[column].dtype

            # Xác định dtype_category của DataFrame
            if pd.api.types.is_numeric_dtype(dfColumnDtype) or pd.api.types.is_float_dtype(dfColumnDtype):
                dtype_category = 'numeric'
            elif pd.api.types.is_datetime64_any_dtype(dfColumnDtype):
                dtype_category = 'datetime64'
            elif pd.api.types.is_bool_dtype(dfColumnDtype):
                dtype_category = 'bool'
            else:
                dtype_category = 'str'

            # Tạo khóa (key) để tìm hàm trong dtype_conversion_map
            key = (dbDataType, dtype_category)

            # Lấy hàm từ dtype_conversion_map, nếu không có thì trả về None
            conversion_func = dtype_conversion_map.get(key,None)

            try:
                if conversion_func:
                    dfImp[column] = conversion_func(dfImp[column], precision, scale,max_length)
                else:
                    dfImp[column] = self.__convert_dtype(dfImp[column], dbDataType)
                
            except Exception as e:
                raise ValueError(f'Error converting row {index}, column {column} [{dfColumnDtype},{dbDataType}]: {e}')

        return dfImp
    # end correctDataType

    def __convert_dtype(self,df_column: pd.Series,dbDataType: pd.Series) -> pd.Series:
        """
        Chuyển đổi kiểu dữ liệu của một cột trong DataFrame và kiểm tra bộ nhớ sử dụng.

        Chuyển đổi kiểu dữ liệu của một cột trong DataFrame.

        Parameters:
        df_column (pd.Series): Cột của DataFrame cần chuyển đổi.

        Returns:
        pd.Series: Cột đã được chuyển đổi kiểu dữ liệu.
        """
        dtype_mapping = {
            'NUMBER': 'float64',
            'VARCHAR2': 'str',
            'DATE': 'datetime64',
            'BOOLEAN': 'bool'
        }   
        dtype = dtype_mapping.get(dbDataType, 'object')
        return df_column.astype(dtype, errors='ignore')
    # end __convert_dtype

    def _convert_bool_to_value(self,df_col,data_type):
        conversion_map = {
            '1': 1, '0': 0,
            'yes': 1, 'no': 0,
            'true': 1, 'false': 0,
            'Yes': 1, 'No': 0,
            'True': 1, 'False': 0
        }
        # Use pandas' replace method to map values to 1 or 0
        converted_df = df_col.str.lower().replace(conversion_map).fillna(0)


        # Nếu data_type là 'NUMBER', trả về dạng số nguyên
        if data_type == 'NUMBER':
            return converted_df.astype(int, errors='ignore')

        # Nếu data_type là 'VARCHAR2', trả về dạng chuỗi '1' hoặc '0'
        elif data_type == 'VARCHAR2':
            return converted_df.astype(str, errors='ignore')

        # Trường hợp khác (nếu có), trả về dạng chuỗi theo mặc định
        return converted_df
    # end _convert_to_bool_numeric


    def __convert_to_datetime(self, df_column: pd.Series) -> pd.Series:
        """
        Chuyển đổi cột từ các định dạng số nguyên hoặc chuỗi khác nhau sang datetime64.

        Parameters:
        df_column (pd.Series): DataFrame chứa cột cần chuyển đổi.

        Returns:
        pd.Series: Column đã được chuyển đổi kiểu dữ liệu.
        """
        
        def detect_format(value):
            formats = {8: '%Y%m%d', 6: '%y%m%d',12:'%Y%m%d%H%M', 14: '%Y%m%d%H%M%S'}
            return formats.get(len(str(int(value))), None)  # Sử dụng int(value) để tránh chuỗi có dấu chấm thập phân
        
        def convert_excel_date(excel_date):
            """
            Chuyển đổi giá trị Excel serial date thành datetime.
            """
            base_date = pd.Timestamp('1899-12-30')
            return base_date + pd.to_timedelta(excel_date, unit='D')
        def regex_convert(value):
            patterns = [
                # Long date formats with time
                (r'(\d{4})/(\d{2})/(\d{2})\s+(\d{2}):(\d{2}):(\d{2})', '%Y/%m/%d %H:%M:%S'),  # YYYY/MM/DD HH:MM:SS
                (r'(\d{4})-(\d{2})-(\d{2})\s+(\d{2}):(\d{2}):(\d{2})', '%Y-%m-%d %H:%M:%S'),  # YYYY-MM-DD HH:MM:SS
                (r'(\d{2})/(\d{2})/(\d{4})\s+(\d{2}):(\d{2}):(\d{2})', '%d/%m/%Y %H:%M:%S'),  # DD/MM/YYYY HH:MM:SS
                (r'(\d{2})-(\d{2})-(\d{4})\s+(\d{2}):(\d{2}):(\d{2})', '%d-%m-%Y %H:%M:%S'),  # DD-MM-YYYY HH:MM:SS
                (r'(\d{2})/(\d{2})/(\d{4})\s+(\d{2}):(\d{2})', '%m/%d/%Y %H:%M'),            # MM/DD/YYYY HH:MM
                (r'(\d{2})-(\d{2})-(\d{4})\s+(\d{2}):(\d{2})', '%m-%d-%Y %H:%M'),            # MM-DD-YYYY HH:MM
                
                # Short date formats
                (r'(\d{4})/(\d{2})/(\d{2})', '%Y/%m/%d'),  # YYYY/MM/DD
                (r'(\d{4})-(\d{2})-(\d{2})', '%Y-%m-%d'),  # YYYY-MM-DD
                (r'(\d{4})(\d{2})(\d{2})', '%Y%m%d'),      # YYYYMMDD
                (r'(\d{2})/(\d{2})/(\d{4})', '%d/%m/%Y'),  # DD/MM/YYYY
                (r'(\d{2})-(\d{2})-(\d{4})', '%d-%m-%Y'),  # DD-MM-YYYY
                (r'(\d{2})/(\d{2})/(\d{4})', '%m/%d/%Y'),  # MM/DD/YYYY
                (r'(\d{2})-(\d{2})-(\d{4})', '%m-%d-%Y'),  # MM-DD-YYYY

                # Full date with day name (e.g., Tuesday, September 24, 2024)
                (r'\w+,\s*(\w+)\s*(\d{1,2}),\s*(\d{4})', '%A, %B %d, %Y'),  # Day, Month DD, YYYY
                (r'\w+,\s*(\d{1,2})\s*(\w+),\s*(\d{4})', '%A, %d %B, %Y'),  # Day, DD Month, YYYY

                # Short date with time and AM/PM (e.g., 9/24/2024 1:21:27 PM)
                (r'(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2}):(\d{2})\s+(AM|PM)', '%m/%d/%Y %I:%M:%S %p'),

                # Short date formats (MM/DD/YYYY)
                (r'(\d{1,2})/(\d{1,2})/(\d{4})', '%m/%d/%Y'),

                # ISO 8601 formats
                (r'(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2}):(\d{2})', '%Y-%m-%dT%H:%M:%S'),  # ISO 8601: YYYY-MM-DDTHH:MM:SS
                (r'(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})', '%Y-%m-%dT%H:%M'),            # ISO 8601: YYYY-MM-DDTHH:MM
                (r'(\d{4})-(\d{2})-(\d{2})\s*(\d{2}):(\d{2})', '%Y-%m-%d %H:%M'),          # YYYY-MM-DD HH:MM (no seconds)

                # Variations without time
                (r'(\d{2})\s+(\w+)\s+(\d{4})', '%d %B %Y'),  # DD Month YYYY (e.g., 15 September 2024)
                (r'(\d{2})\s+(\w+)\s+(\d{2})', '%d %b %y')   # DD Mon YY (e.g., 15 Sep 24)
            ]
            
            # Duyệt qua các mẫu regex và kiểm tra giá trị có khớp với định dạng nào không
            for pattern, date_format in patterns:
                match = re.match(pattern, value)
                if match:
                    try:
                        # Chuyển đổi thành datetime theo định dạng tương ứng
                        return pd.to_datetime(value, format=date_format)
                    except ValueError:
                        return pd.NaT  # Trả về NaT nếu giá trị không hợp lệ
            
            # Trả về NaT nếu không khớp với bất kỳ định dạng nào
            return pd.NaT
        def clean_value(value):
            # Xóa khoảng trắng, ký tự không cần thiết
            if isinstance(value, str):
                value = value.strip()
            return value

        def convert_value(value):
            if pd.isna(value):
                return pd.NaT  # Giữ nguyên giá trị null
            try:
                # Kiểm tra nếu là chuỗi với định dạng "YYYY/MM/DD HH:MM:SS"
                if isinstance(value, str):
                    value = clean_value(value)
                    formats = ['%Y/%m/%d %H:%M:%S',
                               '%Y/%m/%d %H:%M', 
                               '%Y-%m-%d %H:%M:%S',
                               '%Y-%m-%d %H:%M', 
                               '%Y%m%d', 
                               '%Y%m%d%H%M%S', 
                               '%Y%m%d%H%M', 
                               '%y%m%d']
        
                    for fmt in formats:
                        try:
                            return pd.to_datetime(value, format=fmt, errors='raise')
                        except (ValueError, TypeError):
                            continue
                    # end for
                    return regex_convert(value)
                # end value is str
                if isinstance(value,int):
                    format_type = detect_format(value)
                    if format_type:
                        return pd.to_datetime(value, format=format_type, errors='coerce')
                    else:
                        # Nếu không xác định được định dạng, cố gắng chuyển đổi thành datetime mà không có định dạng cụ thể
                        return pd.to_datetime(value, errors='coerce')
                # end value is int
                # Kiểm tra nếu là số thực (Excel serial date)
                if isinstance(value, (float, int)) and value > 59:  # Excel serial date lớn hơn 59 (Excel bug năm 1900)
                    return convert_excel_date(value)
                # end value is float excel serial
                
 
                return pd.NaT
            except (ValueError, TypeError):
                raise ValueError(f"Unsupported date format or value: {value}")

        # Áp dụng hàm convert_value cho từng giá trị trong cột
        return df_column.apply(convert_value)


    def __convert_number_dtype(self, df_column:pd.Series, precision:int, scale:int)->pd.Series:
        # Xử lý giá trị NaN trước khi chuyển đổi
        if df_column.isnull().any():
            df_column = df_column.fillna(0)  

        if scale > 0:
            max_value = 10**(precision - scale) - 1
            min_value = -max_value
            df_column = df_column.round(scale)
            df_column = df_column.apply(lambda x: self.__check_max_value(x, max_value, min_value))
        else:
            max_value = 10**precision - 1
            df_column = df_column.astype(int, errors='ignore')
            df_column = df_column.apply(lambda x: self.__check_max_value(x, max_value))
    
        return df_column
    # end __convert_number_dtype

    def __convert_number_to_str(self, df_column: pd.Series, max_length: int) -> pd.Series:
        # Kiểm tra và thay thế giá trị null bằng chuỗi rỗng
        df_column = df_column.fillna('')

        # Kiểm tra phần thập phân, nếu phần thập phân là 0 thì cắt bỏ và chuyển đổi thành chuỗi
        def convert_value(x):
            if isinstance(x, float) and x.is_integer():
                return str(int(x))
            elif isinstance(x, str) and x.endswith('.0'):
                return x[:-2]
            return str(x)

        # Áp dụng hàm chuyển đổi cho từng giá trị trong cột
        df_column = df_column.apply(convert_value)

        # Cắt chuỗi nếu vượt quá độ dài cho phép
        df_column = df_column.str.slice(0, max_length)
        
        return df_column
    # end __convert_number_to_str

    def convert_str_to_number(self,df_column:pd.Series, precision:int, scale:int, allow_decimal: bool = True) -> pd.Series:
        """
        Chuyển đổi chuỗi thành số nếu giá trị chỉ chứa các ký tự số từ 0-9.
        Kiểm tra và chuyển đổi thập phân thành float nếu được phép.

        Parameters:
        df_column (pd.Series): cột chứa Chuỗi cần chuyển đổi.
        allow_decimal (bool): Cho phép chuyển đổi thành float nếu là thập phân.

        Returns:
        int/float: Giá trị sau khi chuyển đổi thành số nguyên hoặc số thập phân.
        None: Nếu chuỗi không hợp lệ hoặc không phải số.
        """
        def convert(value,allow_decimal=True):
            if re.fullmatch(r'\d+(\.\d+)?', value):
                if '.' in value and allow_decimal:
                    return float(value)
                else:
                    return int(float(value))
            else:
                return 0
        return self.__convert_number_dtype(df_column.apply(lambda x: convert(x, allow_decimal=True)),precision,scale)

    
    

    def mappedHeaderCompareTable(self, fileHeader:list, configHeader:dict) -> list:
        """
        So sánh header từ file và config, mapping lại header trong config.
        Nếu header từ file và config không giống nhau, trả về header từ file.

        Args:
            fileHeader (list): Danh sách header từ file.
            configHeader (dict): Mapping header từ file sang header trong database.

        Returns:
            list: Danh sách header đã được mapping.
        """

        # Kiểm tra độ dài của danh sách header
        if len(fileHeader) != len(configHeader):
            return fileHeader

        # Mapping header
        mapped_header = [configHeader.get(header_key, header_key) for header_key in fileHeader]
        
        

        # Kiểm tra xem tất cả các header đều được mapping thành công
        if len(fileHeader) == len(mapped_header):
            return mapped_header

        # Nếu không mapping thành công, trả về header từ file
        return fileHeader
    
    # def findEmptyRow(self, df:pd.DataFrame) -> int:
    #     empty_row_index = df[df.isnull().all(axis=1)].index.min()
    #     return empty_row_index

    def readToEmptyCol(self, df:pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
     
        if self.endColumnNumer is not None and 0 < self.endColumnNumer <= len(df.columns):
            df = df.iloc[:, :self.endColumnNumer + 1]
        elif self.template.header is not None and 0 < len(self.template.header)<= len(df.columns):
            df = df.iloc[:, :len(self.template.header) + 1]
        else:
            # Kiểm tra hàng đầu tiên để tìm cột trống
            empty_col_index = df.columns[df.iloc[0].isnull()].min()
            if pd.notna(empty_col_index):
                df = df.iloc[:, :empty_col_index]

        return df
    # end readToEmptyCol
        
    
    def readToEmptyRow(self, df:pd.DataFrame) -> pd.DataFrame:
        """
        if config read to empty Row = True >> find index of row nearest.
        convert dataFrame
        else: auto remove empty row
        """
        if df.empty:
            return df
        
        
        if self.template.is_read_to_empty_row == True:
            empty_row_index = df[df.isnull().all(axis=1)].index.min()
            if pd.notna(empty_row_index):
                df = df.iloc[:empty_row_index]
        else:
            df = df.dropna(how='all', inplace=True)

        return df
    # end readToEmptyRow

    def replace_none_with_null(self,value):
        if pd.isna(value):  # Kiểm tra nếu giá trị là NaN hoặc None
            return '<NULL>'
        return value
    
    def __readExcelToPd(self,fileName) -> pd.DataFrame:
         # Đọc file Excel với openpyxl
        wb = load_workbook(fileName)
        sheet = wb.active

        # read data 
        data = []
        #  self.config.get(self.templateName, {}).get(configurationConstant.IMPORT_TMP_END_ROW, sheet.max_row)
        for row in sheet.iter_rows(min_row=self.template.start_row,
                                max_row=self.configEndRow(sheet.max_row),
                                min_col=self.template.start_col, 
                                max_col=self.configEndCol(sheet.max_column)):
            row_data = [cell.value for cell in row]
            data.append(row_data)
        
        # convert to DataFrame
        df = pd.DataFrame(data)

        # remove empty row
        df = self.readToEmptyRow(df)
        # read to empty column
        df = self.readToEmptyCol(df)
        # xử lí null value
        df = df.map(self.replace_none_with_null)
        # build header
        header = self.template.header
        if header is None or not header:
            # Nếu header không được xác định trong config, lấy từ file
            header = [cell.value for cell in sheet[1]]

        mappedTableHeader = self.template.column_mapping
        if mappedTableHeader is not None:
            header = self.mappedHeaderCompareTable(header, mappedTableHeader)

        if header is not None:
            header = [item.upper() for item in header]
        df.columns = header

        

        return df
    # end __readExcelToPd

    def __padasReadExcel(self,fileName, sheetName = None) -> pd.DataFrame:
        # Xử lý cấu hình dòng
        skiprows = max(self.template.start_row - 1, 0) if self.template.start_row and self.template.start_row> 0 else 0
        end_row = self.template.end_row
        
        # Xác định số lượng dòng đọc
        if end_row is None or end_row == "" or end_row <= 0:
            nrows = None
        else:
            nrows = end_row - self.template.start_row + 1

        # Xử lý cấu hình cột
        start_col = self.template.start_row - 1 if self.template.start_row and self.template.start_row > 0 else 0
        end_col = self.template.end_col - 1 if self.template.end_col and self.template.end_col > 0 else None
        # build header
        header = self.template.header
        mappedTableHeader = self.template.column_mapping
        if mappedTableHeader is not None:
            header = self.mappedHeaderCompareTable(header, mappedTableHeader)

        if header is not None:
            header = [item.upper() for item in header]

        # Đọc file Excel
        usecols = range(start_col, end_col + 1) if end_col is not None else None
        df = pd.read_excel(fileName,
                        sheet_name=sheetName if sheetName else 0,
                        header=None if header else 0,  # Đọc mà không sử dụng hàng đầu tiên làm tiêu đề
                        skiprows=skiprows,
                        nrows=nrows,
                        usecols=usecols
                    )
        
        if header:
            df.columns = header

        return df

    @property
    def _is_excel_file(self):
        return self.template.file_extension == DataConstant.RPT_EXCEL_FILE_TYPE or self.template.file_extension == DataConstant.RPT_EXCEL_FILE_TYPE_OLD
    @property
    def _is_csv_file(self):
        return self.template.file_extension == DataConstant.RPT_CSV_FILE_TYPE
    @property
    def _is_txt_file(self):
        return self.file_extension == DataConstant.RPT_TEXT_FILE_TYPE

    def read_file_to_df(self) -> pd.DataFrame:
        """
        Đọc file dữ liệu với nhiều định dạng, xử lý exception và tối ưu hóa.
        Hỗ trợ template động cho file Excel dựa trên file config.

        input:
            self._fileName >>> tham số input khi khởi tạo (str): Đường dẫn đến file dữ liệu. 
            self._config  >>> tham số input khi khởi tạo dựa vào templateFileName (str): Đường dẫn đến file config.

        Returns:
            pd.DataFrame: DataFrame chứa dữ liệu từ file.

        Raises:
            FileNotFoundError: Nếu file dữ liệu hoặc file config không tồn tại.
            PermissionError: Nếu không có quyền truy cập vào file.
            Exception: Nếu gặp lỗi bất kỳ khác trong quá trình đọc file.
        """
        try:
            if self._is_excel_file:
                
                # convert to DataFrame
                df = self.__readExcelToPd(self.fileName)
                # df = self.__padasReadExcel(self.fileName)

            elif self._is_csv_file:
                df = pd.read_csv(self.fileName, header= None if self.template.header is None else 0)
            elif self._is_txt_file:
                df = pd.read_csv(self.fileName, sep= self._sep, header=None)  # Giả sử file text sử dụng tab làm dấu phân cách
            else:
                raise Exception(f"File format is not support: {self.template.file_extension}")

            # Kiểm tra xem file có trống hay không
            if df.empty:
                raise Exception("[{self.__class__.__name__}][{inspect.currentframe().f_code.co_name}]File Empty.")

            return df

        except FileNotFoundError:
            raise FileNotFoundError(f"[{self.__class__.__name__}][{inspect.currentframe().f_code.co_name}]File is not Exis: {self.fileName}")
        except PermissionError:
            raise PermissionError(f"[{self.__class__.__name__}][{inspect.currentframe().f_code.co_name}]Do not Permission: {self.fileName}")
        except Exception as e:
            raise Exception(f"[{self.__class__.__name__}][{inspect.currentframe().f_code.co_name}]Import data read file: Error on reading file: {e}")
            
        
    # end readFile

    
        
    def convert_sql(self, df: pd.DataFrame) -> str:
        #  """ Mapping column Name and process Data Type"""

        mappedColumns = [col for col in df.columns]
        columns = ", ".join(mappedColumns)
        placeholders = ", ".join([":" + col for col in df.columns])
        sql = f"INSERT INTO {self.template.table_name} ({columns}) VALUES ({placeholders})"

        return sql
    # end sqlScript

    def testSql(self, df):
        
        sql1 = f"INSERT INTO {self.template.table_name} "
        rs = ""
        for index, row in df.iterrows():
            col = ", ".join(df.columns)
            value = "', '".join(map(str, row.values))
            rs += f"{sql1} ({col}) VALUES ('{value}');"
        
        return rs

    

        
        

